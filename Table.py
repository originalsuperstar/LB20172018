import openpyxl
import xlsxwriter

def makeTable():
    #Start excel FILE
    workbook = xlsxwriter.Workbook('Table.xlsx')
    worksheet = workbook.add_worksheet("Table")
    bold = workbook.add_format({'bold': True})
    worksheet.write("A1", "Gene ID NCBI", bold)
    worksheet.write("B1", "Locus tag" ,bold)
    worksheet.write("C1", "Gene name" ,bold)
    worksheet.write("D1", "Strand" ,bold)
    worksheet.write("E1", "Protein ID NCBI" ,bold)
    worksheet.write("F1", "Acession number Uniprot" ,bold)
    worksheet.write("G1", "Protein Name" ,bold)
    worksheet.write("H1", "Protein Full Name",bold)
    worksheet.write("I1", "Sequence length" ,bold)
    worksheet.write("J1", "Location" ,bold)
    worksheet.write("K1", "GO Terms" ,bold)
    worksheet.write("L1", "EC number" ,bold)
    worksheet.write("M1", "Função" ,bold)
    #End excel file
    
    #parse Genebank info
    wb = openpyxl.load_workbook('GeneGBInfo/genesGenbank.xlsx')
    first_sheet = wb.get_sheet_names()[0]
    worksheet_open = wb.get_sheet_by_name(first_sheet)

    for row in range(2,worksheet_open.max_row+1):  
        for column in "ABDEH":  
            cell_name = "{}{}".format(column, row)
            value=worksheet_open[cell_name].value
            if 'A' in column:
                loc='A'+str(row)
                worksheet.write(loc,value)
            if 'B' in column:
                loc='D'+str(row)
                worksheet.write(loc,value)
            if 'D' in column:
                loc='C'+str(row)
                worksheet.write(loc,value)
            if 'E' in column:
                loc='B'+str(row)
                worksheet.write(loc,value)
            if 'H' in column:
                loc='E'+str(row)
                worksheet.write(loc,value)   
    
    #parse uniprot info           
    wb = openpyxl.load_workbook('UniprotInfo/genesUniprot.xlsx')
    first_sheet = wb.get_sheet_names()[0]
    worksheet_open = wb.get_sheet_by_name(first_sheet)

    for row in range(2,worksheet_open.max_row+1):  
        for column in "ABCDEHIJ":  
            cell_name = "{}{}".format(column, row)
            value=worksheet_open[cell_name].value
            if 'A' in column:
                loc='F'+str(row)
                worksheet.write(loc,value)
            if 'B' in column:
                loc='G'+str(row)
                worksheet.write(loc,value)
            if 'C' in column:
                loc='H'+str(row)
                worksheet.write(loc,value)
            if 'D' in column:
                loc='L'+str(row)
                worksheet.write(loc,value)
            if 'E' in column:
                loc='M'+str(row)
                worksheet.write(loc,value)
            if 'H' in column:
                loc='J'+str(row)
                worksheet.write(loc,value)
            if 'I' in column:
                loc='K'+str(row)
                worksheet.write(loc,value)  
            if 'J' in column:
                loc='I'+str(row)
                worksheet.write(loc,value)
        
makeTable()